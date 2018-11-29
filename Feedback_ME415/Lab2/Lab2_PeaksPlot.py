'''
ME415 Lab 2 - Gather M Values and phase lag data from excel frequency plot data
'''

import os
from scipy.signal import find_peaks
from openpyxl import load_workbook
import numpy as np, matplotlib.pyplot as plt

freqs = []
outFileName = 'M_values_new2.csv'
with open(outFileName, 'w') as out_file:
    out_file.write('Freq(Hz),M-Value,phase_lag,\n')


for file_name in os.listdir('.'):
    if file_name.endswith('Hz.xlsx'):
        if file_name.startswith('freq_'):
            wkbk = load_workbook(file_name)
            # wkst = wkbk.active
            wkst = wkbk['Sheet1']
            t_dat = []
            in_dat = []
            out_dat = []
            for row in range(2,wkst.max_row+1):
                if row<2502:
                    cell_name = f'D{row}'
                    t_dat.append(wkst[cell_name].value)
                cell_name = f'F{row}'
                in_dat.append(wkst[cell_name].value)
                cell_name = f'G{row}'
                out_dat.append(wkst[cell_name].value)
            new_t = [i-4 for i in t_dat if i<8.002]

            t_dat = new_t + t_dat
            peak_in,_ = find_peaks(in_dat)
            peak_in_vals = [in_dat[i] for i in peak_in if i < 4501]
            peak_in_time = [t_dat[i] for i in peak_in if i < 4501]
            peak_out,_ = find_peaks(out_dat)
            peak_out_vals = [out_dat[i] for i in peak_out if i < 4501]
            peak_out_time = [t_dat[i] for i in peak_out if i < 4501]
            phase_lag = np.mean([peak_in_time[i]-peak_out_time[i] for i in range(len(peak_out_time))])
            print(phase_lag)
            M_val = np.mean(peak_out_vals)/np.mean(peak_in_vals)
            outString = f'{file_name[5:-7]},{M_val},{phase_lag}\n'
            with open(outFileName, 'a') as out_file:
                out_file.write(outString)
            # plt.plot(t_dat[0:4501],in_dat,label='Input')
            # plt.plot(t_dat[0:4501],out_dat,label='Output')
            # plt.plot(t_dat,in_dat[:-1],label='Input')
            # plt.plot(t_dat,out_dat[:-1],label='Output')
            # plt.title(f'{file_name[5:-5]} Data')
            # plt.scatter(peak_in_time,peak_in_vals)
            # plt.scatter(peak_out_time,peak_out_vals)
            # plt.show()